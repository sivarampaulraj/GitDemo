import openpyxl
path = "C:\\Users\\sivar\\Desktop\\PYTHON\\Excel_Python\\Student_Details.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = 1, column = 1)
print(cell_obj.value)
print(wb_obj.sheetnames)
sheet_obj = wb_obj["siva"]
cell_obj = sheet_obj.cell(row = 1, column = 1)
print(cell_obj.value)

for sheet in wb_obj:
    cell_obj = sheet.cell(row = 1, column = 1)
    print(sheet.title)
    print(cell_obj.value)

print("*************")

sheet_obj = wb_obj["Sheet1"]
for x in range(1,5):
    print("-----------")
    for y in range(1,5):
        cell_obj = sheet_obj.cell(row = x, column = y)
        print(cell_obj.value)
        
